import json
import logging
import os

import pandas as pd
from openpyxl import load_workbook

from utils import get_current_month, get_current_day, get_current_month_for_file

logger = logging.getLogger(__name__)


def get_categories(excel_path):
    # TODO: Add error handling and logging
    sheet_name = get_current_month()
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    days = get_payments_table(df)
    return list(days.index)


def get_payments_table(df, range_rows=range(12, 21), range_cols=range(1, 32)):
    try:
        indices = list(df.iloc[range_rows, 0])
        orig_column_names = list(df.iloc[11, range_cols].index)
        column_names = list(df.iloc[11, range_cols].apply(lambda x: int(x)))

        dic = {}
        for old, new in zip(orig_column_names, column_names):
            dic[old] = new

        days = df.iloc[range_rows, range_cols]
        days = days.rename(columns=dic)
        days.index = indices
        days = days.fillna(0.0)
        logging.info('Successfully extracted day by day payments table')
        return days
    except Exception as e:
        logging.error(f"Exception trying to get payments table: {e}")
        return


def write_to_excel(excel_path, sheet_name, df, range_rows, range_cols):
    try:
        workbook = load_workbook(filename=excel_path)
        sheet = workbook[sheet_name]

        for i in range_rows:
            for j in range_cols:
                sheet.cell(i + 2, j + 1, value=df.iloc[i, j])
        workbook.save(filename=excel_path)
        logging.info(f'Successfully saved new workbook at {excel_path}')
        return True

    except Exception as e:
        # TODO: Better exception handling. Everywhere
        logging.error(f"Exception while writing data to Excel, {e}")
        return


def add_to_excel(excel_path, dic):
    sheet_name = get_current_month()
    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        logging.info('Successfully read Excel file')

        range_rows = range(12, 23)
        range_cols = range(1, 32)
        days = get_payments_table(df, range_rows, range_cols)
        if days is None:
            return

        today = int(get_current_day())
        dic = json.loads(dic)
        dic = {k.lower(): v for k, v in dic.items()}

        for item in dic['items']:
            subdic = {k.lower(): v for k, v in item.items()}
            # TODO: get data from receipt, instead of just `today`
            if type(subdic['price']) is str:
                days.loc[subdic['category'], today] += float(subdic['price'].replace('£', ''))
            else:
                days.loc[subdic['category'], today] += float(subdic['price'])

        df.iloc[range_rows, range_cols] = days
        if write_to_excel(excel_path, sheet_name, df, range_rows, range_cols):
            return True
        return

    except Exception as e:
        logging.error(f"Couldn't add data to excel: {e}")
        return
